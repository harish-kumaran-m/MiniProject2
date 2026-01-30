import serial
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os

excel_file = r"rfid_data_log1.xlsx"

# Load or create the Excel workbook and worksheet
if os.path.exists(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Timestamp", "Product ID", "Product Name", "Quantity", "Value", "Tag Timestamp"])

# Function to check if a record exists
def record_exists(parts):
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        if row[1:] == tuple(parts):  # Compare Product ID, Name, Quantity, Value, Tag Timestamp
            return row  # Return the existing row if found
    return None

# Initialize serial connection
ser = serial.Serial('COM10', 9600, timeout=1)
print("Listening on COM10 at 9600 baud...")

while True:
    line = ser.readline().decode('utf-8').strip()
    if line:
        print("Received:", line)
        try:
            parts = line.split(',')
            if len(parts) == 5:
                existing_record = record_exists(parts)
                if existing_record:
                    print("Existing record found:", existing_record)
                    for row in ws.iter_rows(min_row=2):
                        if row[1].value == existing_record[1] and \
                           row[2].value == existing_record[2] and \
                           row[3].value == existing_record[3] and \
                           row[4].value == existing_record[4]:
                            ws.delete_rows(row[0].row)
                            print(f"Existing record deleted at row {row[0].row}.")

                system_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
                ws.append([system_time] + parts)
                wb.save(excel_file)
                print("Data saved to Excel.")
        except Exception as e:
            print("Error:", e)
